import re
from pathlib import Path

from openpyxl import load_workbook

from cicerone.db.connection import get_connection

ROOT = Path(__file__).parent.parent.parent
RESOURCES = ROOT / "resources"
MATRICE_XLSX = RESOURCES / "MatriceDB.xlsx"
CRITERI_MD = RESOURCES / "Criteri_Readiness_Maturity.md"
SCHEMA_SQL = Path(__file__).parent / "schema.sql"

SHEET_READINESS = "AI Readiness-Maturity"
SHEET_NOMI = ["readiness", "implementation"]

RIGA_HEADER = 5
COL_NUM_CRITERIO = 1
COL_NOME_CRITERIO = 2
COL_PRIMO_FRAMEWORK = 5
N_CRITERI = 10


def _normalizza(s: str) -> str:
    return s.replace("’", "'").replace("‘", "'")


def _parse_criteri_md(path: Path) -> list[tuple[str, str]]:
    testo = path.read_text(encoding="utf-8")
    blocchi = re.split(r"^## \d+\.\s+", testo, flags=re.MULTILINE)[1:]
    risultato = []
    for blocco in blocchi:
        righe = blocco.strip().split("\n", 1)
        nome = _normalizza(righe[0].strip())
        definizione = _normalizza(righe[1].strip()) if len(righe) > 1 else ""
        risultato.append((nome, definizione))
    return risultato


def _is_vuoto(conn) -> bool:
    n = conn.execute("SELECT COUNT(*) FROM Criterio").fetchone()[0]
    return n == 0


def _ha_sheet_atteso(conn) -> bool:
    """True se il DB contiene lo sheet 'readiness' atteso dal codice corrente.
    Un DB seedato da una versione con nome sheet diverso (es. 'rediness') ne è
    privo → va ricostruito."""
    return conn.execute(
        "SELECT 1 FROM Sheet WHERE nome = 'readiness'"
    ).fetchone() is not None


def _reset_db(conn) -> None:
    """Ricrea il DB da zero: droppa tutte le tabelle e riapplica lo schema.
    Usato quando il DB persistente è di una versione incompatibile (perdita dei
    dati assessment pregressi accettabile: i dati di seed sono deterministici e
    la coerenza dello schema ha priorità)."""
    conn.execute("PRAGMA foreign_keys = OFF")
    tabelle = [
        r[0]
        for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            "AND name NOT LIKE 'sqlite_%'"
        ).fetchall()
    ]
    for t in tabelle:
        conn.execute(f'DROP TABLE IF EXISTS "{t}"')
    conn.executescript(SCHEMA_SQL.read_text(encoding="utf-8"))
    conn.execute("PRAGMA foreign_keys = ON")


def _seed_tutto(conn) -> None:
    _seed_sheet(conn)
    sheet_id_readiness = conn.execute(
        "SELECT idSheet FROM Sheet WHERE nome = 'readiness'"
    ).fetchone()[0]
    _seed_criteri(conn, sheet_id_readiness)
    _seed_framework_e_voti(conn, sheet_id_readiness)


def run_if_needed() -> None:
    with get_connection() as conn:
        if _is_vuoto(conn):
            _seed_tutto(conn)
            conn.commit()
            return
        # DB già popolato: se manca lo sheet atteso è di una versione
        # incompatibile → ricostruzione + re-seed.
        if not _ha_sheet_atteso(conn):
            _reset_db(conn)
            _seed_tutto(conn)
            conn.commit()


def _seed_sheet(conn) -> None:
    conn.executemany(
        "INSERT INTO Sheet (nome) VALUES (?)",
        [(nome,) for nome in SHEET_NOMI],
    )


def _seed_criteri(conn, sheet_id: int) -> None:
    criteri = _parse_criteri_md(CRITERI_MD)
    if len(criteri) != N_CRITERI:
        raise ValueError(
            f"Attesi {N_CRITERI} criteri in {CRITERI_MD.name}, trovati {len(criteri)}"
        )
    conn.executemany(
        "INSERT INTO Criterio (nomeCriterio, definizione, sheet_id) VALUES (?, ?, ?)",
        [(nome, defn, sheet_id) for nome, defn in criteri],
    )


def _seed_framework_e_voti(conn, sheet_id: int) -> None:
    wb = load_workbook(MATRICE_XLSX, data_only=True)
    ws = wb[SHEET_READINESS]

    nomi_framework: list[tuple[int, str]] = []
    for col in range(COL_PRIMO_FRAMEWORK, ws.max_column + 1):
        nome = ws.cell(row=RIGA_HEADER, column=col).value
        if nome:
            nomi_framework.append((col, str(nome).strip()))

    framework_id_per_col: dict[int, int] = {}
    for col, nome in nomi_framework:
        cur = conn.execute(
            "INSERT INTO Framework (nomeFramework, sheet_id) VALUES (?, ?)",
            (nome, sheet_id),
        )
        framework_id_per_col[col] = cur.lastrowid

    criteri_id_per_numero = {
        i + 1: row[0]
        for i, row in enumerate(
            conn.execute(
                "SELECT idCriterio FROM Criterio WHERE sheet_id = ? ORDER BY idCriterio",
                (sheet_id,),
            ).fetchall()
        )
    }

    voti = []
    riga_primo_criterio = RIGA_HEADER + 1
    for offset in range(N_CRITERI):
        numero = offset + 1
        riga = riga_primo_criterio + offset
        criterio_id = criteri_id_per_numero[numero]
        for col, framework_id in framework_id_per_col.items():
            valore = ws.cell(row=riga, column=col).value
            if valore is None:
                continue
            voto = float(valore)
            if voto not in (0, 2.5, 5, 7.5, 10):
                raise ValueError(
                    f"Voto fuori scala in ({riga},{col}): {voto}"
                )
            voti.append((framework_id, criterio_id, voto))

    conn.executemany(
        "INSERT INTO Voto (framework_id, criterio_id, voto) VALUES (?, ?, ?)",
        voti,
    )


if __name__ == "__main__":
    run_if_needed()
    print("Seed completato.")
